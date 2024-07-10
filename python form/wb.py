from openpyxl import Workbook

# Create a new workbook
wb = Workbook()

# Select the active sheet
ws = wb.active

# Rename the active sheet (optional)
ws.title = "MySheet"

# Write data to the sheet
ws['A1'] = "Hello"
ws['B1'] = "World"
ws['A2'] = 42
ws['B2'] = 3.14

# Save the workbook to a file
file_path = "D:/python form/new_excel_file.xlsx"
wb.save(file_path)

print(f"Excel file created and saved to {file_path}")
